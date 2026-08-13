from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/data-converter/docs/solidworks-costing/Costing-Startwerte-DE-2026.xlsx"

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10)
GREEN = Font(name=FONT, size=10, color="008000")
HEAD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, size=12, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")
HEADFILL = PatternFill("solid", fgColor="1F3864")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EUR = '#,##0.00 "€"'
EUR3 = '#,##0.000 "€"'
NUM0 = '#,##0'
NUM1 = '#,##0.0'
NUM2 = '#,##0.00'
PCT = '0.0%'

wb = Workbook()


def sheet(name, title, subtitle=None):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE
    ws.sheet_view.showGridLines = False
    return ws


def header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD
        c.fill = HEADFILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def body(ws, start_row, rows, formats, fonts=None):
    for r, data in enumerate(rows, start=start_row):
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = BOX
            fmt = formats[i - 1]
            if fmt:
                c.number_format = fmt
            if fonts and fonts[i - 1]:
                c.font = fonts[i - 1]
            elif isinstance(val, str) and val.startswith("="):
                c.font = BLACK
            else:
                c.font = BLUE if fmt else BLACK
    return start_row + len(rows)


def note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE
    return row + 1


# ---------------------------------------------------------------- 00 Anleitung
ws = wb.active
ws.title = "00_Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 105

rows = [
    ("SOLIDWORKS Costing – Startwerte Deutschland 2026", ""),
    ("", ""),
    ("Zweck", "Startdatensatz zum Befüllen leerer Costing-Vorlagen (.sldctm / .sldcts / .sldctc). "
              "Alle Werte sind marktübliche Richtwerte und müssen an eure Kostenstellenrechnung angeglichen werden."),
    ("So einspielen", "1) Costing-Vorlagen-Editor öffnen  2) Vorlage > 'Export nach Excel'  "
                      "3) Werte aus dieser Mappe in die exportierte Datei übertragen (Spalten zuordnen)  "
                      "4) 'Import' zurück in die Vorlage  5) speichern und mit Referenzteilen kalibrieren."),
    ("Wichtig", "Diese Datei ist NICHT direkt importierbar – SOLIDWORKS erwartet seine eigene Spaltenstruktur, "
                "die sich je Version und Sprachstand unterscheidet. Immer erst exportieren, dann füllen."),
    ("Was nicht per Excel geht", "Maschinenstundensätze, Rüstzeiten und benutzerdefinierte Operationen werden im "
                                 "Vorlagen-Editor eingetragen (Blätter 20, 21, 50 dieser Mappe)."),
    ("", ""),
    ("LEGENDE", ""),
    ("blaue Zahl", "Eingabewert – hier anpassen"),
    ("schwarze Zahl", "Formel – nicht überschreiben"),
    ("gelbe Zelle", "zentrale Annahme, die viele Ergebnisse beeinflusst"),
    ("Hinweis zu Formeln", "Die berechneten Spalten (schwarz) werden beim ersten Öffnen in Excel automatisch "
                           "berechnet; in Vorschau-Tools können sie vorher leer erscheinen."),
    ("", ""),
    ("BLÄTTER", ""),
    ("10_Material_Zerspanung", "Werkstoffe, Preis pro kg, Dichte, Schrotterlös – für die Zerspanungsvorlage"),
    ("11_Material_Blech", "Werkstoff × Blechdicke, Tafelformate, Tafelpreise – für die Blechvorlage"),
    ("20_Maschinen_Stundensaetze", "Stundensätze je Verfahren, Basis für alle abgeleiteten Kosten"),
    ("21_Ruestzeiten", "Rüstzeiten je Los und daraus berechnete Rüstkosten"),
    ("30_Zerspanung_MRR", "Zeitspanvolumen und Bohrvorschübe je Materialklasse"),
    ("40_Laser_Schnittwerte", "Schnittgeschwindigkeiten, Einstechzeiten, Schnittkosten je Meter"),
    ("41_Biegen", "Biegezeiten und Kosten je Biegung"),
    ("50_Custom_Operationen", "Oberfläche, Prüfung, Verpackung – Register 'Benutzerdefiniert'"),
    ("60_Aufschlaege", "Ausschuss, Gemeinkosten, Gewinn"),
    ("", ""),
    ("Stand", "August 2026"),
    ("Quellen", "siehe README.md im selben Ordner (Kapitel 7)"),
]
for r, (a, b) in enumerate(rows, start=1):
    ws.cell(row=r, column=1, value=a).font = TITLE if r == 1 else (
        Font(name=FONT, size=10, bold=True) if b and r > 2 else Font(name=FONT, size=10, bold=True))
    c = ws.cell(row=r, column=2, value=b)
    c.font = BLACK
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws["A9"].font = BLUE
ws["A10"].font = BLACK
ws["A11"].fill = YELLOW

# ------------------------------------------------- 10 Material Zerspanung
ws = sheet("10_Material_Zerspanung", "Materialdaten für die Zerspanungsvorlage (.sldctm)",
           "Preise = Handelsware, Klein-/Mittelmengen inkl. Zuschnittaufschlag, Stand 2026. "
           "Werkstoffname muss exakt dem Namen in eurer SOLIDWORKS-Werkstoffdatenbank entsprechen.")
header(ws, 4,
       ["Materialklasse", "Werkstoff (Name in SOLIDWORKS)", "Dichte [kg/m³]",
        "Preis min [€/kg]", "Preis max [€/kg]", "Startwert [€/kg]",
        "Schrotterlös [€/kg]", "Preis je dm³ [€]", "Bemerkung"],
       [18, 30, 14, 12, 12, 13, 13, 13, 46])

mat = [
    ("Stahl", "S235JR", 7850, 1.40, 1.90, 1.65, 0.15, "Standard-Baustahl, Flach/Blech"),
    ("Stahl", "S355J2", 7850, 1.60, 2.10, 1.80, 0.15, "höherfest, Schweißkonstruktionen"),
    ("Stahl", "DC01", 7850, 1.50, 2.00, 1.75, 0.15, "kaltgewalzt, Feinblech"),
    ("Stahl", "DX51D+Z (verzinkt)", 7850, 1.60, 2.20, 1.90, 0.12, "Zinkauflage im Preis enthalten"),
    ("Stahl", "C45 / 1.0503", 7850, 1.70, 2.30, 1.95, 0.15, "Wellen, Drehteile"),
    ("Stahl", "11SMn30 (Automatenstahl)", 7850, 1.90, 2.60, 2.20, 0.15, "gute Zerspanbarkeit"),
    ("Stahl", "42CrMo4", 7850, 2.20, 3.00, 2.50, 0.15, "vergütet, Rundmaterial"),
    ("Stahl", "16MnCr5", 7850, 2.10, 2.90, 2.40, 0.15, "Einsatzstahl"),
    ("Edelstahl", "1.4301 (X5CrNi18-10)", 7900, 8.00, 13.00, 10.50, 1.20, "V2A, Legierungszuschlag monatlich prüfen"),
    ("Edelstahl", "1.4404 (X2CrNiMo17-12-2)", 8000, 10.00, 15.00, 12.50, 1.40, "V4A, seewasserbeständig"),
    ("Edelstahl", "1.4571", 8000, 11.00, 16.00, 13.50, 1.40, "titanstabilisiert"),
    ("Aluminium", "EN AW-5754 (AlMg3)", 2660, 5.50, 8.00, 6.50, 1.10, "Blech, gut umformbar/schweißbar"),
    ("Aluminium", "EN AW-6082 (AlMgSi1)", 2700, 6.00, 9.00, 7.00, 1.10, "Standard für Frästeile"),
    ("Aluminium", "EN AW-7075", 2810, 9.00, 14.00, 11.00, 1.10, "hochfest"),
    ("Aluminium", "EN AW-5083", 2660, 6.50, 9.50, 7.50, 1.10, "seewasserfest"),
    ("Buntmetall", "CuZn39Pb3 (Messing)", 8470, 9.00, 14.00, 11.00, 5.00, "Automatendrehteile"),
    ("Buntmetall", "Cu-ETP (Kupfer)", 8930, 12.00, 17.00, 14.00, 7.00, "Stromschienen"),
    ("Buntmetall", "CuSn8 (Bronze)", 8800, 14.00, 20.00, 16.50, 7.00, "Gleitlager"),
    ("Kunststoff", "POM-C", 1410, 6.00, 10.00, 8.00, 0.00, "Präzisionsteile"),
    ("Kunststoff", "PA6", 1140, 7.00, 11.00, 9.00, 0.00, "Zahnräder, Gleitelemente"),
    ("Kunststoff", "PE-HD", 950, 4.00, 6.00, 5.00, 0.00, "Verschleißleisten"),
    ("Kunststoff", "PMMA", 1190, 6.00, 9.00, 7.50, 0.00, "Sichtscheiben"),
    ("Guss", "EN-GJL-250 (GG25)", 7200, 2.20, 3.20, 2.70, 0.15, "Gehäuse"),
    ("Guss", "EN-GJS-400-15 (GGG40)", 7100, 2.60, 3.80, 3.10, 0.15, "Sphäroguss"),
    ("Titan", "Ti Grade 5 (TiAl6V4)", 4430, 45.00, 85.00, 60.00, 6.00, "Luftfahrt, sehr teuer"),
]
r = 5
for m in mat:
    ws.cell(row=r, column=1, value=m[0]).font = BLACK
    ws.cell(row=r, column=2, value=m[1]).font = BLACK
    for col, val, fmt in ((3, m[2], NUM0), (4, m[3], EUR), (5, m[4], EUR), (6, m[5], EUR), (7, m[6], EUR)):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = fmt
        c.font = BLUE
    c = ws.cell(row=r, column=8, value=f"=C{r}*F{r}/1000")
    c.number_format = EUR
    c.font = BLACK
    ws.cell(row=r, column=9, value=m[7]).font = NOTE
    for col in range(1, 10):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Startwert = empfohlener Eintrag in der Costing-Vorlage. Preis je dm³ dient nur zur Plausibilitätsprüfung.")
note(ws, r + 2, "Bei 1.4301/1.4404 macht der Legierungszuschlag bis zu 40 % des Preises aus – monatlich nachziehen.")

# ------------------------------------------------- 11 Material Blech
ws = sheet("11_Material_Blech", "Materialdaten für die Blechvorlage (.sldcts)",
           "Jede Kombination aus Werkstoff und Dicke ist in der Vorlage eine eigene Zeile. "
           "Fehlt eine Dicke, wird das Teil nicht bewertet.")
header(ws, 4,
       ["Werkstoff", "Dicke [mm]", "Dichte [kg/m³]", "Preis [€/kg]", "Tafel Länge [mm]",
        "Tafel Breite [mm]", "Tafelgewicht [kg]", "Tafelpreis [€]", "Preis je m² [€]", "Schrotterlös [€/kg]"],
       [24, 10, 13, 12, 14, 14, 14, 13, 13, 14])

blech = []
for name, dens, price, scrap, dicken, tafel in [
    ("S235JR", 7850, 1.65, 0.15, [1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0, 8.0, 10.0, 12.0, 15.0, 20.0], (3000, 1500)),
    ("DC01", 7850, 1.75, 0.15, [0.75, 1.0, 1.5, 2.0, 3.0], (2000, 1000)),
    ("DX51D+Z", 7850, 1.90, 0.12, [0.75, 1.0, 1.5, 2.0, 3.0], (2000, 1000)),
    ("1.4301", 7900, 10.50, 1.20, [1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0, 8.0, 10.0, 12.0], (2000, 1000)),
    ("1.4404", 8000, 12.50, 1.40, [1.5, 2.0, 3.0, 4.0, 5.0, 6.0], (2000, 1000)),
    ("EN AW-5754 (AlMg3)", 2660, 6.50, 1.10, [1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0, 8.0, 10.0], (2000, 1000)),
]:
    for d in dicken:
        blech.append((name, d, dens, price, tafel[0], tafel[1], scrap))

r = 5
for b in blech:
    ws.cell(row=r, column=1, value=b[0]).font = BLACK
    for col, val, fmt, f in ((2, b[1], NUM1, BLUE), (3, b[2], NUM0, BLUE), (4, b[3], EUR, BLUE),
                             (5, b[4], NUM0, BLUE), (6, b[5], NUM0, BLUE)):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = fmt
        c.font = f
    ws.cell(row=r, column=7, value=f"=E{r}/1000*F{r}/1000*B{r}/1000*C{r}").number_format = NUM2
    ws.cell(row=r, column=8, value=f"=G{r}*D{r}").number_format = EUR
    ws.cell(row=r, column=9, value=f"=H{r}/(E{r}/1000*F{r}/1000)").number_format = EUR
    c = ws.cell(row=r, column=10, value=b[6])
    c.number_format = EUR
    c.font = BLUE
    for col in (7, 8, 9):
        ws.cell(row=r, column=col).font = BLACK
    for col in range(1, 11):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Tafelgewicht/Tafelpreis sind berechnet – in der Vorlage werden je nach Version Tafelformat und €/kg eingetragen.")
note(ws, r + 2, "Verschnitt bildet Costing über die Schachtelung der Blechtafel ab; hier keinen Verschnittzuschlag einrechnen.")

# ------------------------------------------------- 20 Stundensätze
ws = sheet("20_Maschinen_Stundensaetze", "Maschinenstundensätze Deutschland 2026",
           "inkl. Bediener, Energie, Werkzeugverschleiß und Abschreibung – ohne Gewinn und Verwaltungsgemeinkosten.")
header(ws, 4, ["Verfahren / Maschine", "Spanne min [€/h]", "Spanne max [€/h]", "Startwert [€/h]",
               "Startwert [€/min]", "Hinweis"],
       [34, 15, 15, 15, 15, 52])
saetze = [
    ("Bearbeitungszentrum 3-Achs", 70, 100, 85, "Standard-Frästeile"),
    ("Bearbeitungszentrum 5-Achs", 100, 150, 115, "Komplexteile, weniger Aufspannungen"),
    ("CNC-Drehen 2-Achs", 60, 90, 75, "Futter-/Wellenteile"),
    ("CNC-Drehen mit angetriebenen Werkzeugen", 85, 115, 95, "Komplettbearbeitung"),
    ("Langdrehautomat", 70, 100, 85, "Serienkleinteile"),
    ("Säge / Trennen", 40, 60, 50, "Rohteilzuschnitt"),
    ("Faserlaser 4–6 kW", 130, 180, 150, "Streuung am Markt 40–300 €/h, Mittelwert ~137 €/h"),
    ("Stanz-/Nibbelmaschine", 90, 130, 110, "nur bei vorhandenem Werkzeugsatz"),
    ("Abkantpresse", 70, 95, 80, "Basis für alle Biegekosten"),
    ("Entgraten / Schleifen manuell", 45, 65, 55, "Handarbeitsplatz"),
    ("Schweißen MAG/WIG", 55, 80, 65, "inkl. Zusatzwerkstoff-Umlage"),
    ("Montage / Handarbeitsplatz", 45, 65, 55, "Baugruppenmontage"),
    ("Messraum / Prüfung", 70, 95, 80, "3D-Messmaschine inkl. Bediener"),
]
r = 5
for s in saetze:
    ws.cell(row=r, column=1, value=s[0]).font = BLACK
    for col, val in ((2, s[1]), (3, s[2]), (4, s[3])):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = EUR
        c.font = BLUE
    ws.cell(row=r, column=4).fill = YELLOW
    c = ws.cell(row=r, column=5, value=f"=D{r}/60")
    c.number_format = EUR3
    c.font = BLACK
    ws.cell(row=r, column=6, value=s[4]).font = NOTE
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Gelb = zentrale Annahme. Diese Werte zuerst mit der eigenen Kostenrechnung abgleichen.")
note(ws, r + 2, "Maschinen in der Vorlage mit den realen Anlagennamen benennen (z. B. 'DMU 50', 'TruLaser 3030').")

RATE = "'20_Maschinen_Stundensaetze'"

# ------------------------------------------------- 21 Rüstzeiten
ws = sheet("21_Ruestzeiten", "Rüstzeiten und Rüstkosten je Los",
           "Rüstkosten werden in Costing über die Losgröße auf das Einzelteil verteilt.")
header(ws, 4, ["Vorgang", "Bezug", "Zeit [min]", "Stundensatz [€/h]", "Rüstkosten [€]", "Hinweis"],
       [36, 20, 12, 16, 14, 46])
ruest = [
    ("Fräsen – je Aufspannung", "je Aufspannung", 45, 85, "3-Achs-BAZ, inkl. Nullpunkt und Erstteil"),
    ("Fräsen – zusätzlicher Werkzeugwechsel", "je Werkzeug", 3, 85, "nur bei Werkzeugen außerhalb des Standardsatzes"),
    ("Drehen – je Aufspannung", "je Aufspannung", 30, 75, "inkl. Backenwechsel"),
    ("Sägen – Rohteilzuschnitt", "je Los", 10, 50, ""),
    ("Laserschneiden – je Programm/Tafel", "je Tafel", 10, 150, "Schachtelung und Materialwechsel"),
    ("Stanzen – Werkzeugsatz rüsten", "je Los", 25, 110, ""),
    ("Abkanten – Grundrüsten", "je Los", 15, 80, ""),
    ("Abkanten – je zusätzliches Biegewerkzeug", "je Werkzeug", 6, 80, ""),
    ("Schweißvorrichtung einrichten", "je Los", 20, 65, ""),
    ("Pulverbeschichtung – Los anmelden/aufhängen", "je Los", 20, 55, "zzgl. Mindestlospreis des Beschichters"),
    ("Messmittel/Prüfplan einrichten", "je Los", 15, 80, ""),
]
r = 5
for x in ruest:
    ws.cell(row=r, column=1, value=x[0]).font = BLACK
    ws.cell(row=r, column=2, value=x[1]).font = BLACK
    c = ws.cell(row=r, column=3, value=x[2]); c.number_format = NUM0; c.font = BLUE
    c = ws.cell(row=r, column=4, value=x[3]); c.number_format = EUR; c.font = BLUE
    c = ws.cell(row=r, column=5, value=f"=C{r}/60*D{r}"); c.number_format = EUR; c.font = BLACK
    ws.cell(row=r, column=6, value=x[4]).font = NOTE
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Stundensätze aus Blatt 20 übernommen – bei Änderung dort hier nachziehen.")

# ------------------------------------------------- 30 MRR
ws = sheet("30_Zerspanung_MRR", "Zeitspanvolumen und Bohrvorschübe für die Zerspanungsvorlage",
           "Nicht das Spitzen-MRR des Werkzeugherstellers eintragen, sondern den effektiven Wert "
           "inkl. Luftschnitten, Werkzeugwechseln und Schlichtanteil (Faustregel: Schruppwert × 0,4–0,5).")
header(ws, 4, ["Materialklasse", "Schruppen min [cm³/min]", "Schruppen max [cm³/min]",
               "Schlichten [cm³/min]", "Startwert effektiv [cm³/min]", "Bohrvorschub [mm/min]",
               "Kosten je cm³ bei 85 €/h", "Hinweis"],
       [22, 18, 18, 16, 19, 17, 18, 34])
mrr = [
    ("Aluminium", 150, 400, 30, 120, 500, "6082/5754, HSC-fähig"),
    ("Baustahl / Automatenstahl", 40, 80, 12, 35, 180, "S235, 11SMn30"),
    ("Vergütungsstahl", 25, 50, 8, 22, 140, "42CrMo4, 16MnCr5"),
    ("Edelstahl", 20, 40, 6, 16, 110, "1.4301/1.4404, kühlmittelintensiv"),
    ("Titan", 8, 15, 3, 7, 60, "geringe Schnittgeschwindigkeit"),
    ("Kunststoff", 300, 600, 60, 200, 600, "POM, PA, PE"),
    ("Guss", 50, 90, 14, 40, 200, "GG25, GGG40"),
    ("Buntmetall", 120, 300, 30, 100, 400, "Messing, Bronze"),
]
r = 5
for m in mrr:
    ws.cell(row=r, column=1, value=m[0]).font = BLACK
    for col, val, fmt in ((2, m[1], NUM0), (3, m[2], NUM0), (4, m[3], NUM0), (5, m[4], NUM0), (6, m[5], NUM0)):
        c = ws.cell(row=r, column=col, value=val); c.number_format = fmt; c.font = BLUE
    ws.cell(row=r, column=5).fill = YELLOW
    c = ws.cell(row=r, column=7, value=f"=85/60/E{r}"); c.number_format = EUR3; c.font = BLACK
    ws.cell(row=r, column=8, value=m[6]).font = NOTE
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Spalte G nur zur Plausibilitätsprüfung: Zerspankosten je cm³ abgetragenem Volumen (3-Achs-BAZ 85 €/h).")

# ------------------------------------------------- 40 Laser
ws = sheet("40_Laser_Schnittwerte", "Faserlaser 4–6 kW: Schnittgeschwindigkeiten, Einstechzeiten, Schnittkosten",
           "Bewusst konservativ, weil Costing damit inkl. Beschleunigungs- und Konturverlusten rechnet. "
           "Referenz: 6 kW, 10 mm Edelstahl ≈ 2.000–2.500 mm/min; 20 mm Baustahl ≈ 950 mm/min.")
header(ws, 4, ["Dicke [mm]", "S235 (O₂) [mm/min]", "1.4301 (N₂) [mm/min]", "AlMg3 (N₂) [mm/min]",
               "Einstechzeit [s]", "Schnittkosten S235 [€/m]", "Schnittkosten 1.4301 [€/m]",
               "Kosten je Einstich [€]"],
       [11, 18, 18, 18, 15, 20, 20, 18])
laser = [
    (1.0, 8000, 25000, 20000, 0.3),
    (1.5, 6500, 18000, 14000, 0.3),
    (2.0, 5000, 12000, 9500, 0.3),
    (3.0, 3200, 7000, 5500, 0.3),
    (4.0, 2600, 4800, 3800, 0.8),
    (5.0, 2200, 3600, 2800, 0.8),
    (6.0, 1900, 2900, 2200, 0.8),
    (8.0, 1500, 2100, 1500, 1.5),
    (10.0, 1250, 1600, 1000, 1.5),
    (12.0, 1000, 1150, 700, 2.5),
    (15.0, 800, 800, None, 2.5),
    (20.0, 600, None, None, 4.0),
]
r = 5
for x in laser:
    c = ws.cell(row=r, column=1, value=x[0]); c.number_format = NUM1; c.font = BLUE
    for col, val in ((2, x[1]), (3, x[2]), (4, x[3])):
        c = ws.cell(row=r, column=col, value=val if val is not None else "–")
        if val is not None:
            c.number_format = NUM0
            c.font = BLUE
        else:
            c.font = NOTE
            c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=5, value=x[4]); c.number_format = NUM1; c.font = BLUE
    c = ws.cell(row=r, column=6, value=f"=1000/B{r}*{RATE}!$E$11"); c.number_format = EUR; c.font = BLACK
    if x[2] is not None:
        c = ws.cell(row=r, column=7, value=f"=1000/C{r}*{RATE}!$E$11"); c.number_format = EUR; c.font = BLACK
    else:
        c = ws.cell(row=r, column=7, value="–"); c.font = NOTE; c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=8, value=f"=E{r}/60*{RATE}!$E$11"); c.number_format = EUR3; c.font = BLACK
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Schnittkosten = Schnittzeit × Laser-Minutensatz aus Blatt 20 (Zeile 'Faserlaser 4–6 kW').")
note(ws, r + 2, "Stanzen alternativ: 0,04–0,08 € je Standardhub, 0,15–0,30 € je Umformwerkzeug (Durchzug, Kieme).")

# ------------------------------------------------- 41 Biegen
ws = sheet("41_Biegen", "Biegen / Abkanten – Zeiten und Kosten je Biegung",
           "Kalkulationsbasis Abkantpresse 80 €/h. Mengeneffekte über die Losgröße abbilden, nicht über kleinere Biegepreise.")
header(ws, 4, ["Teileklasse", "Kantenlänge", "Blechdicke", "Zeit je Biegung [min]",
               "Kosten je Biegung [€]", "Handling je Teil [min]", "Handling je Teil [€]", "Hinweis"],
       [22, 18, 16, 17, 17, 17, 17, 40])
biegen = [
    ("klein", "< 500 mm", "1–3 mm", 0.15, 0.20, "Einmannbedienung"),
    ("mittel", "500–1.500 mm", "1–4 mm", 0.26, 0.30, "Standardfall"),
    ("groß", "1.500–2.000 mm", "2–6 mm", 0.45, 0.45, "Anlagehilfe nötig"),
    ("sehr groß / 2-Mann", "> 2.000 mm", "3–8 mm", 0.70, 0.80, "zwei Bediener, Kosten verdoppeln sich faktisch"),
    ("Sonderfall", "enger Radius / schlecht zugänglich", "beliebig", 0.40, 0.40, "Aufschlag 30–60 % auf Standardzeit"),
    ("Hem / Falz", "< 1.000 mm", "1–2 mm", 0.35, 0.30, "zwei Arbeitsgänge"),
]
r = 5
for x in biegen:
    ws.cell(row=r, column=1, value=x[0]).font = BLACK
    ws.cell(row=r, column=2, value=x[1]).font = BLACK
    ws.cell(row=r, column=3, value=x[2]).font = BLACK
    c = ws.cell(row=r, column=4, value=x[3]); c.number_format = NUM2; c.font = BLUE
    c = ws.cell(row=r, column=5, value=f"=D{r}*{RATE}!$E$13"); c.number_format = EUR; c.font = BLACK
    c = ws.cell(row=r, column=6, value=x[4]); c.number_format = NUM2; c.font = BLUE
    c = ws.cell(row=r, column=7, value=f"=F{r}*{RATE}!$E$13"); c.number_format = EUR; c.font = BLACK
    ws.cell(row=r, column=8, value=x[5]).font = NOTE
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Ab Losgröße 50–100 sinken die Stückkosten je Biegung praxisüblich um 20–40 % – das entsteht in Costing "
                "automatisch durch die Verteilung der Rüstkosten.")

# ------------------------------------------------- 50 Custom
ws = sheet("50_Custom_Operationen", "Benutzerdefinierte Operationen (Register 'Benutzerdefiniert')",
           "Kosten ohne Geometriebezug. In der Vorlage an eine Materialklasse oder an 'alle Materialien' hängen.")
header(ws, 4, ["Operation", "Einheit", "min", "max", "Startwert", "Mindestlospreis [€]", "Hinweis"],
       [30, 16, 12, 12, 14, 18, 44])
custom = [
    ("Pulverbeschichten", "€/m²", 22, 35, 28, 55, "RAL-Standard, einfarbig"),
    ("Nasslackieren", "€/m²", 30, 55, 40, 80, "2K, inkl. Grundierung"),
    ("Eloxieren natur E6/EV1", "€/m²", 25, 45, 32, 60, "nur Aluminium"),
    ("Verzinken galvanisch", "€/kg", 0.80, 1.50, 1.10, 30, "blau-/gelbchromatiert"),
    ("Feuerverzinken", "€/kg", 0.55, 0.95, 0.75, 60, "ab ca. 3 mm Blechdicke"),
    ("Brünieren", "€/kg", 0.60, 1.10, 0.85, 30, ""),
    ("Entgraten manuell", "€/Teil", 0.45, 2.75, 1.00, 0, "0,5–3 min bei 55 €/h"),
    ("Gleitschleifen", "€/kg", 0.80, 1.60, 1.10, 40, "Serienteile"),
    ("Gewinde M4–M12 schneiden", "€/Gewinde", 0.10, 0.30, 0.18, 0, "auf der Maschine"),
    ("Senkung", "€/Senkung", 0.05, 0.12, 0.08, 0, ""),
    ("Einpressmutter/-bolzen setzen", "€/Stück", 0.25, 0.55, 0.40, 0, "inkl. Normteil"),
    ("Lasergravur / Beschriftung", "€/Teil", 0.50, 2.00, 1.00, 0, "Typenschildinhalt"),
    ("Wärmebehandlung Vergüten", "€/kg", 1.20, 2.50, 1.70, 80, ""),
    ("Erstmusterprüfbericht (EMPB)", "€/Los", 60, 150, 95, 0, "nur bei Erstteil"),
    ("Endprüfung / Messprotokoll", "€/Teil", 2.00, 8.00, 4.00, 0, "abhängig von Merkmalanzahl"),
    ("Verpackung", "€/Teil", 1.50, 6.00, 2.50, 0, "Kartonage/Palette"),
]
r = 5
for x in custom:
    ws.cell(row=r, column=1, value=x[0]).font = BLACK
    ws.cell(row=r, column=2, value=x[1]).font = BLACK
    for col, val in ((3, x[2]), (4, x[3]), (5, x[4]), (6, x[5])):
        c = ws.cell(row=r, column=col, value=val); c.number_format = EUR; c.font = BLUE
    ws.cell(row=r, column=5).fill = YELLOW
    ws.cell(row=r, column=7, value=x[6]).font = NOTE
    for col in range(1, 8):
        ws.cell(row=r, column=col).border = BOX
    r += 1
note(ws, r + 1, "Oberflächenkosten je m² rechnet Costing über die vom Modell erkannte Oberfläche – Mindestlospreise "
                "sind in Costing nicht direkt abbildbar und gehören in den Aufschlag oder ins Angebot.")

# ------------------------------------------------- 60 Aufschläge
ws = sheet("60_Aufschlaege", "Aufschläge und Kalkulationszuschläge",
           "In Costing im Task-Fenster je Teil als Aufschlag (positiv) bzw. Rabatt (negativ) gesetzt.")
header(ws, 4, ["Position", "Wert", "Bezug", "Hinweis"], [34, 12, 26, 60])
auf = [
    ("Ausschuss / Nacharbeit", 0.02, "auf Herstellkosten", "1–3 %, je nach Prozessstabilität"),
    ("Werkzeugverschleiß", 0.00, "im Stundensatz enthalten", "nicht doppelt aufschlagen"),
    ("Verwaltungs-/Vertriebsgemeinkosten", 0.12, "auf Herstellkosten", "8–15 % je nach Overhead"),
    ("Gewinn Einzelteil/Prototyp", 0.14, "auf Selbstkosten", "12–15 %"),
    ("Gewinn Serie", 0.06, "auf Selbstkosten", "5–8 %"),
    ("Materialteuerungszuschlag", 0.03, "auf Materialkosten", "nur bei langen Angebotsbindungen"),
    ("Mengenrabatt ab 100 Stück", -0.08, "auf Gesamtkosten", "als negativer Aufschlag eintragen"),
]
r = 5
for x in auf:
    ws.cell(row=r, column=1, value=x[0]).font = BLACK
    c = ws.cell(row=r, column=2, value=x[1]); c.number_format = PCT; c.font = BLUE; c.fill = YELLOW
    ws.cell(row=r, column=3, value=x[2]).font = BLACK
    ws.cell(row=r, column=4, value=x[3]).font = NOTE
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = BOX
    r += 1
r += 1
ws.cell(row=r, column=1, value="Beispielrechnung Einzelteil").font = Font(name=FONT, size=10, bold=True)
r += 1
bsp = [("Herstellkosten (aus Costing)", 100.00, None),
       ("+ Ausschuss", None, "=B{h}*B5"),
       ("+ Verwaltung/Vertrieb", None, "=B{h}*B7"),
       ("= Selbstkosten", None, "=B{h}+B{h1}+B{h2}"),
       ("+ Gewinn Einzelteil", None, "=B{s}*B8"),
       ("= Angebotspreis", None, "=B{s}+B{g}")]
base = r
ws.cell(row=base, column=1, value="Herstellkosten (aus Costing)").font = BLACK
c = ws.cell(row=base, column=2, value=100.0); c.number_format = EUR; c.font = BLUE
ws.cell(row=base + 1, column=1, value="+ Ausschuss").font = BLACK
c = ws.cell(row=base + 1, column=2, value=f"=B{base}*B5"); c.number_format = EUR; c.font = BLACK
ws.cell(row=base + 2, column=1, value="+ Verwaltung / Vertrieb").font = BLACK
c = ws.cell(row=base + 2, column=2, value=f"=B{base}*B7"); c.number_format = EUR; c.font = BLACK
ws.cell(row=base + 3, column=1, value="= Selbstkosten").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(row=base + 3, column=2, value=f"=SUM(B{base}:B{base+2})"); c.number_format = EUR; c.font = Font(name=FONT, size=10, bold=True)
ws.cell(row=base + 4, column=1, value="+ Gewinn Einzelteil").font = BLACK
c = ws.cell(row=base + 4, column=2, value=f"=B{base+3}*B8"); c.number_format = EUR; c.font = BLACK
ws.cell(row=base + 5, column=1, value="= Angebotspreis").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(row=base + 5, column=2, value=f"=B{base+3}+B{base+4}"); c.number_format = EUR; c.font = Font(name=FONT, size=10, bold=True)
for rr in range(base, base + 6):
    for col in (1, 2):
        ws.cell(row=rr, column=col).border = BOX
note(ws, base + 7, "Die Beispielrechnung zeigt die Reihenfolge – Costing selbst liefert die Herstellkosten in B{}.".format(base))

wb.save(OUT)
print("saved", OUT)
